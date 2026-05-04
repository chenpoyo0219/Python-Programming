from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
import time
import os

# test
bus_url = "https://ebus.tycg.gov.tw/ebus/driving-map/3220"  #https://ebus.tycg.gov.tw/ebus/driving-map/133

start_label = time.strftime("%m%d%H%M", time.localtime())
out_name = f"bus_132/bus_time_{start_label}.xlsx"

options = Options()
options.add_argument("--headless=new")
options.add_argument("--disable-gpu")
options.add_argument("--no-sandbox")
options.add_argument("--disable-dev-shm-usage")
options.add_argument("--window-size=1280,800")
options.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36")
driver = webdriver.Chrome(options=options)
driver.implicitly_wait(10)  

hour = time.strftime("%H", time.localtime())
min = time.strftime("%M", time.localtime())
all_time = 45
wait = WebDriverWait(driver, 15)
tab_id = "routeInfo"       
dir_station = "中壢"          # 往中壢: "中壢" ,  往中央: "中央大學"

for i in range(all_time):
    if i == 0:
        next_tick = time.monotonic()
        print("開始寫入")
    next_tick += 60
    driver.get(bus_url)

    # try:   # 切時刻表
    #     el = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, f"#{tab_id}")))
    #     driver.execute_script("arguments[0].scrollIntoView({block:'center'}); arguments[0].click();", el)
    # except TimeoutException:
    #     pass

    try:
        button = wait.until(EC.element_to_be_clickable((By.XPATH, f"//button[.//p[contains(.,'往') and contains(.,'{dir_station}')]]")))
        driver.execute_script("arguments[0].scrollIntoView({block:'center'}); arguments[0].click();", button)
    except TimeoutException:
        pass
    time.sleep(5) 

    html = driver.page_source
    bus_soup = BeautifulSoup(html, 'html.parser')
    rows = bus_soup.select("div.MuiListItem-root.MuiListItem-button div.MuiGrid-container")
    bus_dict = {"站序":[], "站名":[], "預估到站時間":[], "目前公車位置":[], "時間":[]}
    
    for row in rows:
        item = row.select("div.MuiGrid-item")
        if len(item) < 4:
            continue

        station_number  = item[0].get_text(strip=True)
        station_name = item[1].get_text(strip=True)
        arrive_time  = item[2].get_text(strip=True)
        plate = item[3].get_text(strip=True)   

        pos = "O" if plate else "X"
        
        if pos == "O":
            now_time= datetime.now().strftime("%H:%M")
        else:
            now_time = pd.NA

        bus_dict["站序"].append(station_number)
        bus_dict["站名"].append(station_name)
        bus_dict["預估到站時間"].append(arrive_time)
        bus_dict["目前公車位置"].append(pos)
        bus_dict["時間"].append(now_time)

    bus_dataframe = pd.DataFrame(bus_dict)
    if not os.path.exists(out_name):
        bus_dataframe.to_excel(out_name, index=False, engine="openpyxl", sheet_name="原始資料")
    else:
        wb = load_workbook(out_name)
        ws = wb.active
        startrow = ws.max_row
        wb.close()

        startrow = startrow + 1 
        with pd.ExcelWriter(out_name, engine="openpyxl", mode="a", if_sheet_exists="overlay") as w:
            bus_dataframe.to_excel(w, index=False, header=False, startrow=startrow, sheet_name="原始資料",)
    print(time.strftime("%H%M%S", time.localtime()))    
    print(f"[{i+1}/{all_time}] 已寫入 {out_name}")
    
    if i != all_time-1:
        time.sleep(max(0, next_tick - time.monotonic())) 

    if i == all_time-1:
        all_station = bus_dict["站名"]
        all_number = bus_dict["站序"]
driver.quit()

all_data_dataframe = pd.read_excel(out_name, engine="openpyxl", sheet_name="原始資料")
O_dataframe = all_data_dataframe.dropna(subset=["時間"])
time_data = pd.to_datetime(O_dataframe["時間"], format="%H:%M", errors="coerce")
O_dataframe["時間"] = time_data.dt.hour * 60 + time_data.dt.minute

if not O_dataframe.empty:
    group = (O_dataframe.groupby("站名", as_index=False)
              .agg(平均分鐘=("時間", "mean")))
    group["平均分鐘"] = group["平均分鐘"].round().astype(int)
    group["實際進站時間"] = group["平均分鐘"].apply(lambda m: f"{m//60:02d}:{m%60:02d}")
else:
    group = pd.DataFrame(columns=["站名", "實際進站時間"])

out_average_dict = pd.DataFrame({"站序":all_number, "站名" :all_station})  

out_dataframe = (
    out_average_dict.merge(
        group[["站名", "實際進站時間"]],  
        on="站名",
        how="left"                       
    )
    .loc[:,["站序", "站名", "實際進站時間"]]  
)

with pd.ExcelWriter(out_name, engine="openpyxl", mode="a", if_sheet_exists="replace") as w:
    out_dataframe.to_excel(w, index=False, sheet_name="計算後時間")
print("完成")
