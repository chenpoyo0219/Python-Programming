# /// script
# requires-python = ">=3.10"
# dependencies = [
#   "pandas",
#   "openpyxl",
#   "prophet",
# ]
# ///

"""
公車到站時間預測系統
- 使用 uv 套件管理（uv run bus_forecast.py）
- 每個 bus_time_*.xlsx 檔案建立一個獨立 Prophet 模型
- 同一檔案內：除最後一趟外全部訓練，最後一趟驗證
- 結果記錄至 result.txt
"""

import pandas as pd
from openpyxl import load_workbook
from prophet import Prophet
from datetime import datetime, time, timedelta
import os
import logging
import glob

logging.getLogger("cmdstanpy").setLevel(logging.WARNING)
logging.getLogger("prophet").setLevel(logging.WARNING)

# ──────────────────────────────────────────────────────
# 官方時刻表（基準比較用，沒有的檔案就跳過）
# ──────────────────────────────────────────────────────
OFFICIAL_SCHEDULE = {
    "bus_time_12_final.xlsx": [
        "12:20", "12:20", "12:21", "12:22", "12:24", "12:25", "12:28", "12:29", "12:29",
        "12:32", "12:34", "12:36", "12:37", "12:38", "12:38", "12:39", "12:40", "12:42",
        "12:43", "12:45"
    ],
    "bus_time_13_final.xlsx": [
        "13:20", "13:21", "13:22", "13:23", "13:24", "13:25", "13:26", "13:27",
        "13:28", "13:29", "13:30", "13:31", "13:32", "13:33", "13:34", "13:35",
        "13:36", "13:37", "13:38", "13:45"
    ],
    "bus_time_14_final.xlsx": [
        "14:20", "14:20", "14:21", "14:22", "14:24", "14:26", "14:28", "14:29", "14:30",
        "14:32", "14:34", "14:36", "14:37", "14:38", "14:38", "14:39", "14:40", "14:41",
        "14:43", "14:45"
    ],
    "bus_time_15_final.xlsx": [
        "15:20", "15:20", "15:21", "15:22", "15:24", "15:26", "15:28", "15:29", "15:30",
        "15:32", "15:34", "15:36", "15:37", "15:38", "15:38", "15:39", "15:40", "15:41",
        "15:43", "15:45"
    ],
}


# ──────────────────────────────────────────────────────
# 讀取 clean_bus_data.py 輸出格式
# 欄位：日期(MMDD) | 站名 | 時間，每趟以空行分隔
# ──────────────────────────────────────────────────────
def read_clean_bus_excel(filepath: str, year: int = 2025) -> list:
    df = pd.read_excel(filepath, dtype=str)
    df.columns = ["日期", "站名", "時間"]

    trips, current_trip = [], []

    for _, row in df.iterrows():
        date_val    = str(row["日期"]).strip()    if pd.notna(row["日期"])  else ""
        station_val = str(row["站名"]).strip()    if pd.notna(row["站名"])  else ""
        time_val    = str(row["時間"]).strip()    if pd.notna(row["時間"])  else ""

        if not date_val or date_val == "nan":
            if current_trip:
                trips.append(current_trip)
                current_trip = []
            continue

        try:
            date_str = (
                f"{year}-{int(date_val[:2]):02d}-{int(date_val[2:]):02d}"
                if len(date_val) == 4 else date_val
            )
        except Exception:
            date_str = date_val

        if ":" not in time_val:
            continue

        current_trip.append((date_str, station_val, time_val))

    if current_trip:
        trips.append(current_trip)

    return trips


# ──────────────────────────────────────────────────────
# 也支援原 TAO.py 格式（A=MMDD, B=站名, C=時間）
# ──────────────────────────────────────────────────────
def read_excel_groups(filename: str, sheetname: str = None,
                      start_row: int = 2, group_size: int = 20,
                      gap: int = 1, year: int = 2025):
    if not os.path.exists(filename):
        return None
    wb = load_workbook(filename)
    ws = wb[sheetname] if sheetname else wb.active

    def fmt_date(value):
        val = str(value).strip()
        return f"{year}-{int(val[:2]):02d}-{int(val[2:]):02d}"

    def fmt_time(value):
        if isinstance(value, time):
            return value.strftime('%H:%M')
        return str(value).strip() if value is not None else None

    result, row = [], start_row
    while True:
        if ws[f'A{row}'].value is None or ws[f'C{row}'].value is None:
            break
        for i in range(group_size):
            dc, sc, tc = ws[f'A{row+i}'], ws[f'B{row+i}'], ws[f'C{row+i}']
            if None in (dc.value, sc.value, tc.value):
                break
            result.append((fmt_date(dc.value), str(sc.value).strip(), fmt_time(tc.value)))
        row += group_size + gap
    return result


def groups_to_trips(groups: list, group_size: int = 20) -> list:
    return [groups[i:i + group_size] for i in range(0, len(groups), group_size)]


# ──────────────────────────────────────────────────────
# 建立 Prophet 訓練 DataFrame（帶累積站距權重）
# ──────────────────────────────────────────────────────
def build_train_df(trips: list, group_size: int = 20):
    station_map = {0: "起點 (發車前5分)"}
    diff_collector = {i: [] for i in range(1, group_size + 2)}
    temp_trip_data = []
    start_times = []

    for trip in trips:
        if len(trip) < group_size:
            continue

        orig_start_dt = datetime.strptime(f"{trip[0][0]} {trip[0][2]}", "%Y-%m-%d %H:%M")
        orig_last_dt  = datetime.strptime(f"{trip[-1][0]} {trip[-1][2]}", "%Y-%m-%d %H:%M")
        real_date_dt  = datetime.strptime(trip[0][0], "%Y-%m-%d")
        virtual_0_dt  = orig_start_dt - timedelta(minutes=5)
        start_times.append(virtual_0_dt)

        trip_elapsed = [0.0]
        for i, (d, s, t) in enumerate(trip, start=1):
            station_map[i] = s
            current_dt = datetime.strptime(f"{d} {t}", "%Y-%m-%d %H:%M")
            trip_elapsed.append((current_dt - virtual_0_dt).total_seconds() / 60.0)

        last_idx = group_size + 1
        station_map[last_idx] = "終點 (抵達後5分)"
        trip_elapsed.append(
            (orig_last_dt + timedelta(minutes=5) - virtual_0_dt).total_seconds() / 60.0
        )

        for i in range(1, len(trip_elapsed)):
            diff_collector[i].append(trip_elapsed[i] - trip_elapsed[i - 1])

        temp_trip_data.append((real_date_dt, trip_elapsed))

    avg_diffs = {
        i: (sum(v) / len(v) if v else 1.5)
        for i, v in diff_collector.items()
    }

    weight_map, cumulative = {0: 0.0}, 0.0
    for i in range(1, group_size + 2):
        cumulative += avg_diffs[i]
        weight_map[i] = round(cumulative, 4)

    records = []
    for real_date_dt, elapsed_list in temp_trip_data:
        for i, val in enumerate(elapsed_list):
            records.append({
                "ds": real_date_dt,
                "station_idx": weight_map[i],
                "y": val
            })

    return pd.DataFrame(records), station_map, start_times, weight_map


# ──────────────────────────────────────────────────────
# 訓練 Prophet 並預測
# ──────────────────────────────────────────────────────
def train_and_forecast(df, station_map, start_times, weight_map, target_date_str: str):
    m = Prophet(
        yearly_seasonality=False,
        weekly_seasonality=False,
        daily_seasonality=False,
        changepoint_prior_scale=0.2,
        changepoint_range=0.95
    )
    m.add_regressor('station_idx', prior_scale=2)
    m.fit(df)

    target_date = datetime.strptime(target_date_str, "%Y-%m-%d")
    future_df = pd.DataFrame({
        "ds": [target_date] * len(station_map),
        "station_idx": [weight_map[i] for i in range(len(station_map))]
    })
    forecast = m.predict(future_df)

    target_hour = pd.Series([dt.hour for dt in start_times]).mode()[0]
    target_starts = [dt for dt in start_times if dt.hour == target_hour]
    avg_start_mins = (
        sum(dt.hour * 60 + dt.minute for dt in target_starts) / len(target_starts)
        if target_starts else target_hour * 60 + 15
    )
    base_time = target_date + timedelta(minutes=avg_start_mins)

    results = []
    for i, row in forecast.iterrows():
        elapsed = max(0, row["yhat"])
        results.append({
            "station_idx": i,
            "station_name": station_map[i],
            "predicted_time": (base_time + timedelta(minutes=elapsed)).strftime("%H:%M"),
            "elapsed_mins": round(elapsed, 1),
            "weight_used": round(weight_map[i], 2)
        })

    return pd.DataFrame(results)


# ──────────────────────────────────────────────────────
# 誤差計算
# ──────────────────────────────────────────────────────
def compute_errors(predicted: list, reference: list) -> tuple:
    fmt = "%H:%M"
    p = [datetime.strptime(t, fmt) for t in predicted]
    r = [datetime.strptime(t, fmt) for t in reference]
    if len(p) != len(r):
        raise ValueError(f"長度不一致 (預測:{len(p)} vs 參考:{len(r)})")
    diffs = [abs((a - b).total_seconds()) / 60 for a, b in zip(p, r)]
    return diffs, round(sum(diffs) / len(diffs), 2)


# ──────────────────────────────────────────────────────
# 處理單一檔案
# ──────────────────────────────────────────────────────
def process_file(filepath: str) -> tuple:
    """回傳 (lines, avg_ai, avg_official)"""
    filename = os.path.basename(filepath)
    lines = []
    lines.append("=" * 65)
    lines.append(f"檔案: {filename}")
    lines.append("=" * 65)

    # 讀取 trips（先嘗試 clean 格式，再 fallback）
    trips = read_clean_bus_excel(filepath)
    if not trips:
        groups = read_excel_groups(filepath)
        if groups:
            trips = groups_to_trips(groups)

    if not trips:
        lines.append("❌ 無法讀取資料，略過。")
        lines.append("")
        return lines, None, None

    if len(trips) < 2:
        lines.append(f"⚠️ 只有 {len(trips)} 趟，至少需要 2 趟（1訓練+1驗證），略過。")
        lines.append("")
        return lines, None, None

    # ── 切分訓練 / 驗證 ──
    train_trips = trips[:-1]  # 除最後一趟
    val_trip    = trips[-1]   # 最後一趟

    group_size = len(val_trip)

    lines.append(f"總趟數 : {len(trips)}")
    lines.append(f"訓練趟數: {len(train_trips)}")
    lines.append(f"驗證趟數: 1（最後一趟，日期 {val_trip[0][0]}）")
    lines.append(f"每趟站數: {group_size}")
    lines.append("")

    # ── 建模 & 預測 ──
    df, station_map, start_times, weight_map = build_train_df(train_trips, group_size)
    forecast_df = train_and_forecast(
        df, station_map, start_times, weight_map, val_trip[0][0]
    )

    core_predicted = forecast_df["predicted_time"].tolist()[1:-1]  # 去掉虛擬起終點
    val_actual     = [row[2] for row in val_trip]

    # ── 預測詳情 ──
    lines.append("【預測詳情（含虛擬起終點）】")
    lines.append(f"{'站序':>4}  {'站名':<22}  {'預測時間':>8}  {'耗時(分)':>8}  {'累積權重':>8}")
    lines.append("-" * 62)
    for _, row in forecast_df.iterrows():
        lines.append(
            f"{int(row['station_idx']):>4}  {row['station_name']:<22}  "
            f"{row['predicted_time']:>8}  {row['elapsed_mins']:>8.1f}  {row['weight_used']:>8.2f}"
        )
    lines.append("")

    avg_ai = None
    avg_official = None

    # ── AI 模型 vs 實際 ──
    if len(core_predicted) == len(val_actual):
        diffs_ai, avg_ai = compute_errors(core_predicted, val_actual)
        lines.append("【AI 模型 vs 實際到站】")
        lines.append(f"{'站名':<22}  {'AI預測':>7}  {'實際':>7}  {'誤差(分)':>8}")
        lines.append("-" * 52)
        for i, (pred, actual, diff) in enumerate(zip(core_predicted, val_actual, diffs_ai)):
            sname = station_map.get(i + 1, f"站{i+1}")
            lines.append(f"{sname:<22}  {pred:>7}  {actual:>7}  {diff:>8.1f}")
        lines.append(f"\nAI 平均誤差 : {avg_ai} 分鐘")
        lines.append(f"AI 最大誤差 : {max(diffs_ai)} 分鐘")
        lines.append(f"AI 最小誤差 : {min(diffs_ai)} 分鐘")
        lines.append("")
    else:
        lines.append(
            f"⚠️ 站數不符，無法比較 (預測:{len(core_predicted)} vs 實際:{len(val_actual)})"
        )
        lines.append("")

    # ── 官方時刻表 vs 實際 ──
    official = OFFICIAL_SCHEDULE.get(filename)
    if official:
        if len(val_actual) == len(official):
            diffs_off, avg_official = compute_errors(val_actual, official)
            lines.append("【官方時刻表 vs 實際到站】")
            lines.append(f"{'站名':<22}  {'官方':>7}  {'實際':>7}  {'誤差(分)':>8}")
            lines.append("-" * 52)
            for i, (off, actual, diff) in enumerate(zip(official, val_actual, diffs_off)):
                sname = station_map.get(i + 1, f"站{i+1}")
                lines.append(f"{sname:<22}  {off:>7}  {actual:>7}  {diff:>8.1f}")
            lines.append(f"\n官方 平均誤差 : {avg_official} 分鐘")
            lines.append(f"官方 最大誤差 : {max(diffs_off)} 分鐘")
            lines.append(f"官方 最小誤差 : {min(diffs_off)} 分鐘")
            lines.append("")
        else:
            lines.append("⚠️ 官方時刻表站數不符，略過比較。")
            lines.append("")

    # ── 綜合結論 ──
    if avg_ai is not None and avg_official is not None:
        lines.append("【綜合比較】")
        lines.append(f"{'指標':<20}  {'AI模型':>10}  {'官方時刻表':>12}")
        lines.append("-" * 48)
        lines.append(f"{'平均誤差 (分鐘)':<20}  {avg_ai:>10.2f}  {avg_official:>12.2f}")
        better = "✅ AI 模型較準確" if avg_ai < avg_official else "📋 官方時刻表較準確"
        lines.append(f"\n結論: {better}")
        lines.append("")
    elif avg_ai is not None:
        lines.append(f"（無官方時刻表可比較，AI 平均誤差: {avg_ai} 分鐘）")
        lines.append("")

    return lines, avg_ai, avg_official


# ──────────────────────────────────────────────────────
# 主程式
# ──────────────────────────────────────────────────────
def main():
    all_files = sorted(glob.glob("bus_time_*.xlsx"))

    if not all_files:
        print("❌ 找不到任何 bus_time_*.xlsx 檔案。")
        return

    print(f"找到 {len(all_files)} 個檔案，每個獨立建模...\n")

    all_lines = []
    all_lines.append("=" * 65)
    all_lines.append("公車到站時間預測結果報告")
    all_lines.append(f"產生時間: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    all_lines.append(f"檔案數量: {len(all_files)} 個")
    all_lines.append("策略: 每個檔案獨立建立 Prophet 模型，各自以最後一趟驗證")
    all_lines.append("=" * 65)
    all_lines.append("")

    summary_rows = []

    for filepath in all_files:
        fname = os.path.basename(filepath)
        print(f"  處理 {fname} ...", end=" ", flush=True)
        file_lines, avg_ai, avg_official = process_file(filepath)
        all_lines.extend(file_lines)
        summary_rows.append((fname, avg_ai, avg_official))
        print("完成")

    # ── 全體摘要 ──
    all_lines.append("=" * 65)
    all_lines.append("【全體摘要】")
    all_lines.append("=" * 65)
    all_lines.append(f"{'檔案':<25}  {'AI平均誤差':>11}  {'官方平均誤差':>13}  {'較佳':>8}")
    all_lines.append("-" * 65)
    for fname, ai_err, off_err in summary_rows:
        ai_str  = f"{ai_err:.2f} 分"  if ai_err  is not None else "N/A"
        off_str = f"{off_err:.2f} 分" if off_err is not None else "N/A"
        if ai_err is not None and off_err is not None:
            better = "AI" if ai_err < off_err else "官方"
        else:
            better = "-"
        all_lines.append(f"{fname:<25}  {ai_str:>11}  {off_str:>13}  {better:>8}")

    all_lines.append("")
    all_lines.append("=" * 65)
    all_lines.append("報告結束")
    all_lines.append("=" * 65)

    result_text = "\n".join(all_lines)

    with open("result.txt", "w", encoding="utf-8") as f:
        f.write(result_text)

    print("\n" + result_text)
    print("\n✅ 結果已儲存至 result.txt")


if __name__ == "__main__":
    main()